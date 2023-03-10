"""creating Excel file with contents"""
from openpyxl import Workbook
# creating a list to enter into the Excel data cells
favorite_foods = ['Pizza', 'Chocolate Cake', 'Broccoli']
favorite_color = ['Blue', 'Purple', 'Green', 'Orange']
# making a new workbook
workbook = Workbook()
# indicating the active worksheet that's being worked on
worksheet = workbook.active
# .title will title the current active worksheet
worksheet.title = 'Favorite Things'
# This adds 'Favorite Foods' into the row 1, col 1 cell
# This is only adding an item and not a list so no for loop needed
worksheet.cell(1, 1, 'Favorite Foods')
worksheet.cell(1, 2, 'Favorite Colors')
# creating for loops listing only the favorite food items in first col
for index, food in enumerate(favorite_foods):
    # + 1 is the counter to make sure that the data being entered is moving to the next row of the col
    # 3 arguments here - (row + 1/means move to the next row after data is entered, col, data)
    worksheet.cell(index + 2, 1, food)
for index, color in enumerate(favorite_color):
    worksheet.cell(index + 2, 2, color)
# be sure to save worksheet when finish like when opening and closing files
workbook.save('favorites.xlsx')

